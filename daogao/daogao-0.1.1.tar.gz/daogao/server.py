# -*- coding: utf8 -*-
from __future__ import print_function
from tornado.web import RequestHandler
from tornado.ioloop import IOLoop
import tornado.web
from psycopg2 import connect as pg_connect
from psycopg2.extras import DictCursor
from openpyxl.workbook import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, PatternFill
from openpyxl.cell import Cell
from email.mime.text import MIMEText
from email.utils import formataddr
import smtplib
import requests
import sys
import mandrill
import os
from binascii import hexlify
from threading import Thread
try:
    import simplejson as json
except ImportError:
    import json


hilos = []

unique_id = lambda: hexlify(os.urandom(16))
downloads_url = 'http://daogao.local'

def send_email(to=None, subject=None, content=""):
    api_key = "key-86dcacf19db8cf3716276f66992814d2"
    url_template = "https://api.mailgun.net/v3/{}/messages"
    domain = "edesarrollos.info"
    url = url_template.format(domain)
    response = requests.post(url,
        auth=("api", api_key,),
        data={
            "from": "postmaster@edesarrollos.info",
            "to": to,
            "subject": subject,
            "html": content
            }
        )

def process_xls(data):
    origin = data['dataOrigin']
    header = data['header']
    title = header['title']
    book = Workbook()
    sheet = book.active
    doc_id = unique_id()

    if 'logoURL' in header:
        try:
            response = requests.get(header['logoURL'], stream=True)
            logo = Image(response.raw)
            logo = Image(logo.image.resize((100, 100)))
        except requests.ConnectionError as cerror:
            print(cerror, file=sys.stderr)
            pass
    else:
        logo = None


    hdr_bkg_color = header['backgroundColor']
    header_bkg = PatternFill(fill_type="solid",
                             start_color=hdr_bkg_color,
                             end_color=hdr_bkg_color)

    if origin == 'array':
        rows = data['rows']

        cell = Cell(sheet, value=title)
        cell.alignment = Alignment(horizontal='center',
                                   vertical='center')

        sheet.append(['', '', '', cell])

        sheet.merge_cells('A1:C1')
        sheet.merge_cells('D1:G1')

        for row in rows:
            sheet.append(row)

    else:
        db = data['database']
        sql_query = data['sqlQuery']
        url_callback = data['urlCallback']
        title = data['title']
        columns = data['columns']

        conn = pg_connect(host=db['host'],
                          database=db['name'],
                          password=db['password'],
                          user=db['user'])

        cursor = conn.cursor()
        cursor.execute(sql_query)

        index = 0

        is_first = True
        for row in cursor:
            if is_first:
                sheet.merge_cells('A1:C1')
                sheet.merge_cells('D1:G1')

                sheet.append(['', '', '', cell])

                if logo:
                    sheet.add_image(logo, 'A1')

                headcells = []
                for col in columns:
                    cell = Cell(sheet, value=col['label'])
                    cell.fill = header_bkg
                    headcells.append(cell)

                sheet.append(headcells)

                is_first = False
                #sheet.row_dimensions[0].height = 300
                sheet.row_dimensions[1].height = 72
            
            sheet.append(row)

            index += 1

    """
        files = {'file': open('./archivo.xlsx', 'r')}
        response = requests.post(url_callback, files=files)
        print(response.content)
    """

    book.save('./{}.xlsx'.format(doc_id))
    
    if 'triggers' in data:
        for trigger in data['triggers']:
            trigger_type = trigger['type']
            if trigger_type == 'sendEmail':
                filename = '{}.xlsx'.format(doc_id)
                filepath = '{}/{}'.format(downloads_url, filename)
                content = 'Este es tu archivo: {}'\
                        .format(filepath)

                send_email(to=trigger['emails'],
                           subject='Tu excel está listo',
                           content=content,
                           )

class GenerationHandler(RequestHandler):

    def post(self):
        args = json.loads(self.request.body)
        thr = Thread(target=process_xls, args=(args, ))
        hilos.append(thr)
        thr.start()

        print("hay {} hilos".format(len(hilos)))

        self.write('OK')


urls = [
    (r'^/generate-xls', GenerationHandler),
]

if __name__ == '__main__':
    app = tornado.web.Application(urls)
    loop = IOLoop()
    try:
        app.listen(8081)
        loop.start()
    except KeyboardInterrupt:
        print("bye")